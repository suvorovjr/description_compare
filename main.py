import re
import csv
import zipfile
import openpyxl
from pathlib import Path
from utils import cleanhtml, desc_identificate, get_encoding, create_excel_file


def description_compare(filename):
    # регулярное выражение для очистки строк от имеющейся html разметки
    cleanr = re.compile('<.*?>|&([a-z0-9]+|#[0-9]{1,6}|#x[0-9a-f]{1,6});|_x\\S{1,5}_')

    path = Path(__file__).parent

    path_to_file = path.joinpath(filename)

    # Путь к файлу для записи совпавших id
    save_id = path.joinpath(f"{path_to_file.name.split('.')[0]}_id.xlsx")
    create_excel_file(save_id)

    # Путь для записи совпаших названий (создать заранее)
    save_info = path.joinpath(f"{path_to_file.name.split('.')[0]}_info.xlsx")
    create_excel_file(save_info)

    # Название к zip файлу
    name_zip = path.joinpath(f"{path_to_file.name.split('.')[0]}.zip")

    # Получение кодировки файла
    encoding = get_encoding(path_to_file)

    # Чтение csv файла и его запись в переменную в виде списка
    with open(path_to_file, "r", encoding=encoding) as file:
        reader = csv.reader(file, delimiter=";")
        data = list(reader)

    # номера строк для записи в Excel файлы
    save_id_row = 1
    save_info_row = 1
    # Список id совпавших товаров
    working_id = []
    compare_count = 0

    for i, row in enumerate(data):
        # Проверка строки на валидность
        if i == 0 or len(row) != 5:
            continue
        product_id, product_name, description, product_brand, product_collection = row
        # Проверка есть ли id в списке отработанных id
        if product_id in working_id:
            continue
        identifications_desc = []
        clean_description = cleanhtml(cleanr, description)
        identifications_desc.append(clean_description)
        identification_list = []
        for n, compare_row in enumerate(data):
            if n == 0 or i == n or len(compare_row) != 5:
                continue
            compare_id, compare_name, compare_description, compare_brand, compare_collection = compare_row
            clean_compare_description = cleanhtml(cleanr, compare_description)
            ident_percent_list = []
            if len(product_collection) > 1:
                if product_collection == compare_collection and product_brand == compare_brand:
                    for ident_description in identifications_desc:
                        ident_percent = desc_identificate(ident_description, clean_compare_description)
                        ident_percent_list.append(ident_percent)
                else:
                    continue
            else:
                if product_brand == compare_brand:
                    for ident_description in identifications_desc:
                        ident_percent = desc_identificate(ident_description, clean_compare_description)
                        ident_percent_list.append(ident_percent)
                else:
                    continue
            if max(ident_percent_list) > 80 and compare_id not in working_id:
                print(f"Установлено совпадение: {product_name} == {compare_name} = {max(ident_percent_list)}")
                identifications_desc.append(clean_compare_description)
                identification_list.append([compare_id, compare_name, int(max(ident_percent_list))])
                working_id.append(compare_id)
        if len(identification_list) > 0:
            workbook = openpyxl.load_workbook(save_id)
            sheet = workbook.active
            sheet.cell(row=save_id_row, column=1, value=product_id)
            column = 2
            for value_id in identification_list:
                sheet.cell(row=save_id_row, column=column, value=value_id[0])
                column += 1
            workbook.save(save_id)
            workbook.close()
            info_workbook = openpyxl.load_workbook(save_info)
            info_sheet = info_workbook.active
            info_sheet.cell(row=save_info_row, column=1, value=product_id)
            info_sheet.cell(row=save_info_row, column=2, value=product_name)
            info_sheet.cell(row=save_info_row, column=3, value=clean_description)
            info_sheet.cell(row=save_info_row, column=4, value=product_brand)
            info_sheet.cell(row=save_info_row, column=5, value=product_collection)
            save_info_row += 1
            for values in identification_list:
                compare_id, compare_name, ident_percent = values
                info_sheet.cell(row=save_info_row, column=2, value=compare_name)
                info_sheet.cell(row=save_info_row, column=3, value=compare_id)
                info_sheet.cell(row=save_info_row, column=4, value=ident_percent)
                save_info_row += 1
            info_workbook.save(save_info)
            info_workbook.close()
            print(f"Отработан столбец № {save_id_row}. Установлено совпадений: {len(identification_list)}")
            print(i)
            working_id.append(product_id)
            save_id_row += 1

    with zipfile.ZipFile(name_zip, 'w') as myzip:
        myzip.write(save_id.name)
        myzip.write(save_info.name)
    path_to_file.unlink()
    save_id.unlink()
    save_info.unlink()
